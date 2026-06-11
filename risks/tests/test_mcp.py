"""MCP tests for the risks module."""

import json

import pytest

from accounts.tests.factories import UserFactory
from compliance.tests.factories import ComplianceActionPlanFactory
from mcp.server import McpServer
from mcp.tools import register_all_tools
from risks.models import (
    ISO27005Risk,
    RiskAcceptance,
    RiskTreatmentPlan,
    Threat,
    Vulnerability,
)
from risks.tests.factories import RiskAssessmentFactory, RiskFactory


pytestmark = pytest.mark.django_db


def _call_tool(srv, user, tool_name, arguments):
    result = srv.handle_request(json.dumps({
        "jsonrpc": "2.0",
        "id": 1,
        "method": "tools/call",
        "params": {"name": tool_name, "arguments": arguments},
    }), user)
    raw = result["result"]["content"][0]["text"]
    return json.loads(raw)


class TestRiskAcceptanceApproveMCP:
    def setup_method(self):
        self.srv = McpServer()
        register_all_tools(self.srv)
        self.user = UserFactory(is_superuser=True)

    def _make_acceptance(self):
        risk = RiskFactory()
        return RiskAcceptance.objects.create(
            risk=risk, justification="Approve me", status="active",
        )

    def test_approve_tool_is_registered(self):
        assert "approve_risk_acceptance" in self.srv._tools

    def test_approve_sets_approval_fields(self):
        acceptance = self._make_acceptance()
        assert acceptance.is_approved is False
        result = _call_tool(
            self.srv, self.user, "approve_risk_acceptance",
            {"id": str(acceptance.pk)},
        )
        assert "error" not in result, result
        acceptance.refresh_from_db()
        assert acceptance.is_approved is True
        assert acceptance.approved_by == self.user
        assert acceptance.approved_at is not None

    def test_approve_unknown_id_returns_error(self):
        import uuid
        result = _call_tool(
            self.srv, self.user, "approve_risk_acceptance",
            {"id": str(uuid.uuid4())},
        )
        assert "error" in result


class TestTreatmentPlanActionPlanLinkMCP:
    def setup_method(self):
        self.srv = McpServer()
        register_all_tools(self.srv)
        self.user = UserFactory(is_superuser=True)

    def _make_plan(self):
        risk = RiskFactory()
        return RiskTreatmentPlan.objects.create(
            risk=risk, name="Mitigate", treatment_type="mitigate",
        )

    def test_tools_are_registered(self):
        expected = {
            "list_treatment_plan_action_plans",
            "link_treatment_plan_action_plans",
            "unlink_treatment_plan_action_plans",
            "set_treatment_plan_action_plans",
        }
        assert expected.issubset(self.srv._tools.keys())

    def test_list_empty(self):
        plan = self._make_plan()
        result = _call_tool(
            self.srv, self.user, "list_treatment_plan_action_plans",
            {"treatment_plan_id": str(plan.pk)},
        )
        assert result["total"] == 0
        assert result["items"] == []

    def test_link_and_list(self):
        plan = self._make_plan()
        ap1 = ComplianceActionPlanFactory(is_approved=True)
        ap2 = ComplianceActionPlanFactory(is_approved=True)
        result = _call_tool(
            self.srv, self.user, "link_treatment_plan_action_plans",
            {"treatment_plan_id": str(plan.pk), "action_plan_ids": [str(ap1.pk), str(ap2.pk)]},
        )
        assert result["added"] == 2
        assert result["total"] == 2

        listed = _call_tool(
            self.srv, self.user, "list_treatment_plan_action_plans",
            {"treatment_plan_id": str(plan.pk)},
        )
        assert listed["total"] == 2
        refs = {item["reference"] for item in listed["items"]}
        assert refs == {ap1.reference, ap2.reference}

    def test_link_is_additive(self):
        plan = self._make_plan()
        ap1 = ComplianceActionPlanFactory(is_approved=True)
        ap2 = ComplianceActionPlanFactory(is_approved=True)
        plan.related_action_plans.add(ap1)
        result = _call_tool(
            self.srv, self.user, "link_treatment_plan_action_plans",
            {"treatment_plan_id": str(plan.pk), "action_plan_ids": [str(ap2.pk)]},
        )
        assert result["added"] == 1
        assert result["total"] == 2

    def test_unlink_removes_selected(self):
        plan = self._make_plan()
        ap1 = ComplianceActionPlanFactory(is_approved=True)
        ap2 = ComplianceActionPlanFactory(is_approved=True)
        plan.related_action_plans.add(ap1, ap2)
        result = _call_tool(
            self.srv, self.user, "unlink_treatment_plan_action_plans",
            {"treatment_plan_id": str(plan.pk), "action_plan_ids": [str(ap1.pk)]},
        )
        assert result["removed"] == 1
        assert result["total"] == 1
        assert plan.related_action_plans.first() == ap2

    def test_set_replaces_all(self):
        plan = self._make_plan()
        ap1 = ComplianceActionPlanFactory(is_approved=True)
        ap2 = ComplianceActionPlanFactory(is_approved=True)
        ap3 = ComplianceActionPlanFactory(is_approved=True)
        plan.related_action_plans.add(ap1, ap2)
        result = _call_tool(
            self.srv, self.user, "set_treatment_plan_action_plans",
            {"treatment_plan_id": str(plan.pk), "action_plan_ids": [str(ap3.pk)]},
        )
        assert result["total"] == 1
        assert plan.related_action_plans.first() == ap3

    def test_set_empty_clears(self):
        plan = self._make_plan()
        plan.related_action_plans.add(ComplianceActionPlanFactory())
        result = _call_tool(
            self.srv, self.user, "set_treatment_plan_action_plans",
            {"treatment_plan_id": str(plan.pk), "action_plan_ids": []},
        )
        assert result["total"] == 0

    def test_unknown_treatment_plan_returns_error(self):
        import uuid
        result = _call_tool(
            self.srv, self.user, "list_treatment_plan_action_plans",
            {"treatment_plan_id": str(uuid.uuid4())},
        )
        assert "error" in result

    def test_link_unknown_action_plan_returns_error(self):
        import uuid
        plan = self._make_plan()
        result = _call_tool(
            self.srv, self.user, "link_treatment_plan_action_plans",
            {"treatment_plan_id": str(plan.pk), "action_plan_ids": [str(uuid.uuid4())]},
        )
        assert "error" in result


class TestGenerateRiskRegisterMCP:
    def setup_method(self):
        self.srv = McpServer()
        register_all_tools(self.srv)
        self.user = UserFactory(is_superuser=True)

    def test_tool_is_registered(self):
        assert "generate_risk_register" in self.srv._tools

    def test_generates_report(self):
        from reports.constants import ReportType
        from reports.models import Report
        RiskFactory.create_batch(2)
        result = _call_tool(
            self.srv, self.user, "generate_risk_register", {},
        )
        assert "error" not in result, result
        assert result["report_type"] == ReportType.RISK_REGISTER
        assert result["status"] == "completed"
        assert result["file_name"].endswith(".xlsx")
        # Persisted Report
        report = Report.objects.get(pk=result["id"])
        assert report.file_content
        assert report.file_name.endswith(".xlsx")

class TestRiskAssessmentRiskTreatmentPlanApproveMCP:
    """A5: approve_risk_assessment, approve_risk, approve_risk_treatment_plan."""

    def setup_method(self):
        self.srv = McpServer()
        register_all_tools(self.srv)
        self.user = UserFactory(is_superuser=True)

    def test_approve_tools_registered(self):
        for tool_name in (
            "approve_risk_assessment",
            "approve_risk",
            "approve_risk_treatment_plan",
        ):
            assert tool_name in self.srv._tools

    def test_approve_risk_assessment(self):
        assessment = RiskAssessmentFactory()
        result = _call_tool(
            self.srv, self.user, "approve_risk_assessment",
            {"id": str(assessment.pk)},
        )
        assert "error" not in result, result
        assessment.refresh_from_db()
        assert assessment.is_approved is True
        assert assessment.approved_by == self.user

    def test_approve_risk(self):
        risk = RiskFactory()
        result = _call_tool(
            self.srv, self.user, "approve_risk",
            {"id": str(risk.pk)},
        )
        assert "error" not in result, result
        risk.refresh_from_db()
        assert risk.is_approved is True

    def test_approve_risk_treatment_plan(self):
        risk = RiskFactory()
        plan = RiskTreatmentPlan.objects.create(
            risk=risk, name="Plan-approve", treatment_type="mitigate",
        )
        result = _call_tool(
            self.srv, self.user, "approve_risk_treatment_plan",
            {"id": str(plan.pk)},
        )
        assert "error" not in result, result
        plan.refresh_from_db()
        assert plan.is_approved is True


class TestThreatVulnerabilityIso27005ApproveMCP:
    """B1: approve_threat, approve_vulnerability, approve_iso27005_risk."""

    def setup_method(self):
        self.srv = McpServer()
        register_all_tools(self.srv)
        self.user = UserFactory(is_superuser=True)

    def test_approve_tools_registered(self):
        for tool_name in (
            "approve_threat",
            "approve_vulnerability",
            "approve_iso27005_risk",
        ):
            assert tool_name in self.srv._tools

    def test_approve_threat(self):
        threat = Threat.objects.create(name="ApproveT", type="deliberate")
        result = _call_tool(
            self.srv, self.user, "approve_threat", {"id": str(threat.pk)},
        )
        assert "error" not in result, result
        threat.refresh_from_db()
        assert threat.is_approved is True
        assert threat.approved_by == self.user

    def test_approve_vulnerability(self):
        vuln = Vulnerability.objects.create(name="ApproveV", severity="medium")
        result = _call_tool(
            self.srv, self.user, "approve_vulnerability", {"id": str(vuln.pk)},
        )
        assert "error" not in result, result
        vuln.refresh_from_db()
        assert vuln.is_approved is True

    def test_approve_iso27005_risk(self):
        threat = Threat.objects.create(name="T-mcp", type="deliberate")
        vuln = Vulnerability.objects.create(name="V-mcp", severity="medium")
        analysis = ISO27005Risk.objects.create(
            assessment=RiskAssessmentFactory(),
            threat=threat, vulnerability=vuln,
        )
        result = _call_tool(
            self.srv, self.user, "approve_iso27005_risk",
            {"id": str(analysis.pk)},
        )
        assert "error" not in result, result
        analysis.refresh_from_db()
        assert analysis.is_approved is True


class TestGenerateRiskRegisterMCPFilters:
    """Status filter test, kept separate from the registration class."""

    def setup_method(self):
        self.srv = McpServer()
        register_all_tools(self.srv)
        self.user = UserFactory(is_superuser=True)

    def test_filter_by_status(self):
        RiskFactory(name="StatusKeep", status="analyzed", is_approved=True)
        RiskFactory(name="StatusDrop", status="closed", is_approved=True)
        result = _call_tool(
            self.srv, self.user, "generate_risk_register",
            {"status": "analyzed"},
        )
        assert "error" not in result, result
        from reports.models import Report
        report = Report.objects.get(pk=result["id"])
        content = bytes(report.file_content)
        # Load the workbook and assert on the actual cell values.
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        all_values = []
        for row in ws.iter_rows(min_row=5):
            for c in row:
                if c.value:
                    all_values.append(str(c.value))
        joined = "\n".join(all_values)
        assert "StatusKeep" in joined
        assert "StatusDrop" not in joined
