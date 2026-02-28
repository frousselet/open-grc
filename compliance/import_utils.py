import io
import json

from django.db import transaction
from django.utils.translation import gettext as _
from openpyxl import load_workbook
from openpyxl import Workbook as XlWorkbook
from openpyxl.styles import Alignment, Font, PatternFill

from .constants import (
    FrameworkCategory,
    FrameworkType,
    RequirementCategory,
    RequirementType,
)
from .models import Framework, Requirement, Section


# ── Sample data ──────────────────────────────────────────────

SAMPLE_DATA = {
    "framework": {
        "reference": "EXAMPLE-001",
        "name": "Sample framework",
        "short_name": "Sample",
        "framework_version": "1.0",
        "type": "standard",
        "category": "information_security",
        "issuing_body": "My organization",
        "description": "This is a sample framework to illustrate the import format.",
    },
    "sections": [
        {
            "reference": "SEC.1",
            "name": "Governance",
            "description": "Information security governance measures.",
            "sections": [
                {
                    "reference": "SEC.1.1",
                    "name": "Security policies",
                    "description": "Definition and review of policies.",
                    "requirements": [
                        {
                            "reference": "REQ.1.1.1",
                            "name": "General security policy",
                            "description": "An information security policy must be defined and approved by management.",
                            "guidance": "The policy must be communicated to all staff and reviewed at regular intervals.",
                            "type": "mandatory",
                            "category": "organizational",
                        },
                        {
                            "reference": "REQ.1.1.2",
                            "name": "Policy review",
                            "description": "Security policies must be reviewed at planned intervals.",
                            "guidance": "The review must take place at least once a year or upon significant changes.",
                            "type": "mandatory",
                            "category": "organizational",
                        },
                    ],
                },
            ],
        },
        {
            "reference": "SEC.2",
            "name": "Asset management",
            "description": "Measures related to the management of information assets.",
            "requirements": [
                {
                    "reference": "REQ.2.1",
                    "name": "Asset inventory",
                    "description": "An inventory of information assets must be established and maintained.",
                    "guidance": "The inventory must identify the owner of each asset.",
                    "type": "mandatory",
                    "category": "organizational",
                },
                {
                    "reference": "REQ.2.2",
                    "name": "Information classification",
                    "description": "Information must be classified according to its sensitivity.",
                    "guidance": "",
                    "type": "recommended",
                    "category": "organizational",
                },
            ],
        },
    ],
}


def generate_sample_json():
    """Return a BytesIO containing the sample JSON file."""
    buf = io.BytesIO()
    buf.write(json.dumps(SAMPLE_DATA, ensure_ascii=False, indent=2).encode("utf-8"))
    buf.seek(0)
    return buf


def generate_sample_excel():
    """Return a BytesIO containing the sample Excel (.xlsx) file."""
    wb = XlWorkbook()
    ws = wb.active
    ws.title = "Framework"

    headers = [
        "type", "reference", "name", "description", "guidance",
        "req_type", "req_category",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    fw = SAMPLE_DATA["framework"]
    rows = [
        [
            "framework", fw["reference"], fw["name"], fw["description"], "",
            fw["type"], fw["category"],
        ],
    ]

    def _flatten_sections(sections):
        for sec in sections:
            rows.append([
                "section", sec["reference"], sec["name"],
                sec.get("description", ""), "", "", "",
            ])
            for req in sec.get("requirements", []):
                rows.append([
                    "requirement", req["reference"], req["name"],
                    req.get("description", ""), req.get("guidance", ""),
                    req.get("type", ""), req.get("category", ""),
                ])
            _flatten_sections(sec.get("sections", []))

    _flatten_sections(SAMPLE_DATA["sections"])

    for row_data in rows:
        ws.append(row_data)

    # Auto-width
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value) if cell.value else ""
            max_len = max(max_len, min(len(val), 60))
        ws.column_dimensions[col_letter].width = max_len + 3

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── JSON parser ──────────────────────────────────────────────


def parse_json(file_obj):
    """Parse a JSON file into the intermediate structure."""
    raw = json.load(file_obj)

    fw_data = raw.get("framework", {})
    framework = {
        "reference": fw_data.get("reference", ""),
        "name": fw_data.get("name", ""),
        "short_name": fw_data.get("short_name", ""),
        "description": fw_data.get("description", ""),
        "framework_version": fw_data.get("framework_version", ""),
        "type": fw_data.get("type", ""),
        "category": fw_data.get("category", ""),
        "issuing_body": fw_data.get("issuing_body", ""),
    }

    sections = []
    stats = {"section_count": 0, "requirement_count": 0, "max_depth": 0}

    def _walk_sections(items, parent_ref, depth):
        order_counter = 0
        for item in items:
            order_counter += 1
            section = {
                "reference": item.get("reference", ""),
                "name": item.get("name", ""),
                "description": item.get("description", ""),
                "parent_reference": parent_ref,
                "order": order_counter,
                "depth": depth,
                "requirements": [],
            }
            stats["section_count"] += 1
            if depth > stats["max_depth"]:
                stats["max_depth"] = depth

            req_order = 0
            for req in item.get("requirements", []):
                req_order += 1
                section["requirements"].append({
                    "reference": req.get("reference", ""),
                    "name": req.get("name", ""),
                    "description": req.get("description", ""),
                    "guidance": req.get("guidance", ""),
                    "type": req.get("type", ""),
                    "category": req.get("category", ""),
                    "order": req_order,
                })
                stats["requirement_count"] += 1

            sections.append(section)

            # Recurse into nested sections
            children = item.get("sections", [])
            if children:
                _walk_sections(children, item.get("reference", ""), depth + 1)

    _walk_sections(raw.get("sections", []), None, 1)

    return {
        "framework": framework,
        "sections": sections,
        "stats": stats,
    }


# ── Excel parser ─────────────────────────────────────────────


def parse_excel(file_obj):
    """Parse an Excel (.xlsx) file into the intermediate structure."""
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"framework": {}, "sections": [], "stats": {"section_count": 0, "requirement_count": 0, "max_depth": 0}}

    # Header row
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    col = {name: idx for idx, name in enumerate(headers)}

    def cell(row, name):
        idx = col.get(name)
        if idx is None or idx >= len(row):
            return ""
        val = row[idx]
        return str(val).strip() if val is not None else ""

    framework = {}
    section_refs = {}  # ref -> section_data
    section_list = []  # ordered list of sections
    orphan_requirements = []  # requirements with no matching section

    stats = {"section_count": 0, "requirement_count": 0, "max_depth": 0}

    for row in rows[1:]:
        row_type = cell(row, "type").lower()

        if row_type == "framework":
            framework = {
                "reference": cell(row, "reference"),
                "name": cell(row, "name"),
                "short_name": cell(row, "short_name") if "short_name" in col else "",
                "description": cell(row, "description"),
                "framework_version": cell(row, "framework_version") if "framework_version" in col else "",
                "type": cell(row, "req_type"),
                "category": cell(row, "req_category"),
                "issuing_body": cell(row, "issuing_body") if "issuing_body" in col else "",
            }

        elif row_type == "section":
            ref = cell(row, "reference")
            parent_ref = _find_parent_prefix(ref, section_refs)
            depth = _calc_depth(ref, section_refs, parent_ref)
            if depth > stats["max_depth"]:
                stats["max_depth"] = depth

            section_data = {
                "reference": ref,
                "name": cell(row, "name"),
                "description": cell(row, "description"),
                "parent_reference": parent_ref,
                "order": stats["section_count"] + 1,
                "depth": depth,
                "requirements": [],
            }
            section_refs[ref] = section_data
            section_list.append(section_data)
            stats["section_count"] += 1

        elif row_type == "requirement":
            ref = cell(row, "reference")
            parent_section_ref = _find_parent_prefix(ref, section_refs)

            req_data = {
                "reference": ref,
                "name": cell(row, "name"),
                "description": cell(row, "description"),
                "guidance": cell(row, "guidance"),
                "type": cell(row, "req_type"),
                "category": cell(row, "req_category"),
                "order": 0,  # will be set below
            }

            if parent_section_ref and parent_section_ref in section_refs:
                section = section_refs[parent_section_ref]
                req_data["order"] = len(section["requirements"]) + 1
                section["requirements"].append(req_data)
            else:
                orphan_requirements.append(req_data)

            stats["requirement_count"] += 1

    wb.close()

    # Attach orphan requirements to a virtual root (no section) by adding
    # them as a section-less list. We'll store them as a fake section with
    # empty reference if needed, or just leave them as warnings.
    # For now, create a special section-less entry at the end if orphans exist.
    if orphan_requirements:
        for i, req in enumerate(orphan_requirements, 1):
            req["order"] = i
        section_list.append({
            "reference": "",
            "name": _("(Requirements without section)"),
            "description": "",
            "parent_reference": None,
            "order": stats["section_count"] + 1,
            "depth": 1,
            "requirements": orphan_requirements,
            "is_virtual": True,
        })

    return {
        "framework": framework,
        "sections": section_list,
        "stats": stats,
    }


def _find_parent_prefix(ref, known_refs):
    """Find the longest known reference that is a prefix of `ref`.

    We split by '.' and try progressively shorter prefixes.
    """
    parts = ref.split(".")
    for length in range(len(parts) - 1, 0, -1):
        candidate = ".".join(parts[:length])
        if candidate in known_refs:
            return candidate
    return None


def _calc_depth(ref, known_refs, parent_ref):
    """Calculate depth based on parent chain."""
    depth = 1
    current = parent_ref
    while current and current in known_refs:
        depth += 1
        current = known_refs[current].get("parent_reference")
    return depth


# ── Validation ───────────────────────────────────────────────


def validate_parsed_data(parsed, existing_framework=None):
    """Validate parsed data. Returns (errors, warnings) lists of strings.

    If *existing_framework* is a Framework instance, validation is adapted for
    import-into-existing mode (skip framework uniqueness, check collisions with
    existing sections/requirements).
    """
    errors = []
    warnings = []

    fw = parsed.get("framework", {})

    # Framework required fields — name is always required (used to rename)
    if not existing_framework:
        if not fw.get("reference"):
            errors.append(_("The framework 'reference' field is required."))
    if not fw.get("name"):
        errors.append(_("The framework 'name' field is required."))

    # Framework type/category validation (only relevant for new frameworks)
    if not existing_framework:
        valid_fw_types = {c.value for c in FrameworkType}
        if fw.get("type") and fw["type"] not in valid_fw_types:
            errors.append(
                _("Unknown framework type: '%(type)s'. "
                  "Accepted values: %(values)s") % {
                    "type": fw["type"],
                    "values": ", ".join(sorted(valid_fw_types)),
                }
            )

        valid_fw_categories = {c.value for c in FrameworkCategory}
        if fw.get("category") and fw["category"] not in valid_fw_categories:
            errors.append(
                _("Unknown framework category: '%(category)s'. "
                  "Accepted values: %(values)s") % {
                    "category": fw["category"],
                    "values": ", ".join(sorted(valid_fw_categories)),
                }
            )

        # Check for existing framework with same reference
        if fw.get("reference") and Framework.objects.filter(reference=fw["reference"]).exists():
            errors.append(
                _("A framework with the reference '%(reference)s' already exists in the database.") % {
                    "reference": fw["reference"],
                }
            )

    # Collect existing refs when importing into existing framework
    existing_section_refs = set()
    existing_req_refs = set()
    if existing_framework:
        existing_section_refs = set(
            existing_framework.sections.values_list("reference", flat=True)
        )
        existing_req_refs = set(
            existing_framework.requirements.values_list("reference", flat=True)
        )

    # Section validation
    section_refs = set()
    for sec in parsed.get("sections", []):
        if sec.get("is_virtual"):
            continue
        ref = sec.get("reference", "")
        if not ref:
            errors.append(_("A section has no reference."))
        elif ref in section_refs:
            errors.append(_("Duplicate section reference: '%(ref)s'.") % {"ref": ref})
        else:
            section_refs.add(ref)
            if ref in existing_section_refs:
                errors.append(
                    _("Section '%(ref)s' already exists in framework "
                      "'%(framework_ref)s'.") % {
                        "ref": ref,
                        "framework_ref": existing_framework.reference,
                    }
                )

        if not sec.get("name"):
            errors.append(_("Section '%(ref)s' has no name.") % {"ref": ref})

    # Requirement validation
    valid_req_types = {c.value for c in RequirementType}
    valid_req_categories = {c.value for c in RequirementCategory}
    req_refs = set()

    for sec in parsed.get("sections", []):
        for req in sec.get("requirements", []):
            ref = req.get("reference", "")
            if not ref:
                errors.append(_("A requirement has no reference."))
            elif ref in req_refs:
                errors.append(_("Duplicate requirement reference: '%(ref)s'.") % {"ref": ref})
            else:
                req_refs.add(ref)
                if ref in existing_req_refs:
                    errors.append(
                        _("Requirement '%(ref)s' already exists in framework "
                          "'%(framework_ref)s'.") % {
                            "ref": ref,
                            "framework_ref": existing_framework.reference,
                        }
                    )

            if not req.get("name"):
                errors.append(_("Requirement '%(ref)s' has no name.") % {"ref": ref})

            if req.get("type") and req["type"] not in valid_req_types:
                errors.append(
                    _("Unknown requirement type for '%(ref)s': '%(type)s'. "
                      "Accepted values: %(values)s") % {
                        "ref": ref,
                        "type": req["type"],
                        "values": ", ".join(sorted(valid_req_types)),
                    }
                )

            if req.get("category") and req["category"] not in valid_req_categories:
                errors.append(
                    _("Unknown requirement category for '%(ref)s': '%(category)s'. "
                      "Accepted values: %(values)s") % {
                        "ref": ref,
                        "category": req["category"],
                        "values": ", ".join(sorted(valid_req_categories)),
                    }
                )

        # Warn about orphan sections
        if sec.get("is_virtual"):
            warnings.append(
                _("%(count)s requirement(s) without an identified parent section.") % {
                    "count": len(sec["requirements"]),
                }
            )

    if not parsed.get("sections"):
        warnings.append(_("No sections found in the file."))

    return errors, warnings


# ── Import execution ─────────────────────────────────────────


def execute_import(parsed, owner, created_by, existing_framework=None):
    """Create or update a Framework with Sections and Requirements.

    When *existing_framework* is provided the sections and requirements are
    added to it and its name is overwritten with the value from the file.
    Otherwise a brand-new Framework is created.

    Returns the Framework instance.
    """
    with transaction.atomic():
        fw_data = parsed["framework"]

        if existing_framework:
            framework = existing_framework
            framework.name = fw_data["name"]
            framework.save(update_fields=["name"])
            # Seed section_map with sections already present so that
            # parent_reference resolution works for mixed hierarchies.
            section_map = {
                s.reference: s for s in framework.sections.all()
            }
        else:
            framework = Framework.objects.create(
                reference=fw_data["reference"],
                name=fw_data["name"],
                short_name=fw_data.get("short_name", ""),
                description=fw_data.get("description", ""),
                framework_version=fw_data.get("framework_version", ""),
                type=fw_data.get("type", "other"),
                category=fw_data.get("category", "other"),
                issuing_body=fw_data.get("issuing_body", ""),
                status="draft",
                owner=owner,
                created_by=created_by,
            )
            section_map = {}

        # Create sections (order matters: parents before children)
        for sec_data in parsed["sections"]:
            if sec_data.get("is_virtual"):
                # Virtual section: create requirements without section
                for req_data in sec_data.get("requirements", []):
                    _create_requirement(framework, None, req_data, created_by)
                continue

            parent = None
            if sec_data.get("parent_reference"):
                parent = section_map.get(sec_data["parent_reference"])

            section = Section.objects.create(
                framework=framework,
                parent_section=parent,
                reference=sec_data["reference"],
                name=sec_data["name"],
                description=sec_data.get("description", ""),
                order=sec_data.get("order", 0),
            )
            section_map[sec_data["reference"]] = section

            # Create requirements for this section
            for req_data in sec_data.get("requirements", []):
                _create_requirement(framework, section, req_data, created_by)

    return framework


def _create_requirement(framework, section, req_data, created_by):
    """Create a single Requirement instance."""
    Requirement.objects.create(
        framework=framework,
        section=section,
        reference=req_data["reference"],
        name=req_data["name"],
        description=req_data.get("description", ""),
        guidance=req_data.get("guidance", ""),
        type=req_data.get("type") or "mandatory",
        category=req_data.get("category", ""),
        order=req_data.get("order", 0),
        created_by=created_by,
    )
