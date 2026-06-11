"""PDF report generators."""

import re

from django.template.loader import render_to_string
from django.utils import timezone

from accounts.models import CompanySettings


def _clean_html(text):
    """Strip empty HTML paragraphs and whitespace-only tags."""
    if not text:
        return ""
    # Remove empty paragraphs: <p></p>, <p> </p>, <p>&nbsp;</p>, <p><br></p>
    cleaned = re.sub(r"<p>\s*(?:&nbsp;|<br\s*/?>)?\s*</p>", "", text)
    return cleaned.strip()


def _natural_sort_key(text):
    """Return a sort key that orders numeric parts numerically."""
    parts = re.split(r"(\d+)", text)
    return [int(p) if p.isdigit() else p.casefold() for p in parts]


def build_soa_frameworks_data(frameworks):
    """Build the per-framework data structure consumed by the SoA template.

    Exposed publicly so tests can assert on the structured data without
    invoking weasyprint. Each framework dict carries `framework`, `rows`
    (sorted by requirement number) and `linked_risk_count` (deduplicated).
    """
    from django.db.models import Prefetch

    from compliance.models import ComplianceActionPlan
    from core.workflow import reportable
    from risks.models import Risk

    frameworks_data = []
    for fw in reportable(frameworks):
        requirements = reportable(fw.requirements.all()).select_related(
            "section",
        ).prefetch_related(
            Prefetch("linked_risks", queryset=reportable(Risk.objects.all())),
            Prefetch("action_plans", queryset=reportable(ComplianceActionPlan.objects.all())),
        ).order_by("requirement_number", "created_at")

        rows = []
        framework_risk_ids = set()
        for req in requirements:
            risks = list(req.linked_risks.all())
            risks_data = [
                {
                    "reference": r.reference,
                    "name": r.name,
                    "current_risk_level": r.current_risk_level,
                    "residual_risk_level": r.residual_risk_level,
                    "treatment_decision": (
                        r.get_treatment_decision_display()
                        if r.treatment_decision
                        else ""
                    ),
                    "treatment_decision_key": r.treatment_decision or "",
                    "status": r.get_status_display() if r.status else "",
                }
                for r in risks
            ]
            framework_risk_ids.update(r.pk for r in risks)

            if req.is_applicable:
                plans = req.action_plans.all()
                action_plan_names = [p.name for p in plans]
                justification = ", ".join(action_plan_names)
                # If the requirement is selected primarily to mitigate risks,
                # surface that as the justification when no action plan is
                # named yet.
                if not justification and risks_data:
                    justification = "Selected to address linked risks."
            else:
                justification = req.applicability_justification

            rows.append({
                "number": req.requirement_number or req.reference,
                "name": req.name,
                "is_applicable": req.is_applicable,
                "justification": justification,
                "risks_data": risks_data,
            })

        rows.sort(key=lambda r: _natural_sort_key(r["number"]))

        frameworks_data.append({
            "framework": fw,
            "rows": rows,
            "linked_risk_count": len(framework_risk_ids),
        })
    return frameworks_data


def generate_soa_pdf(frameworks, user):
    """Generate a Statement of Applicability (SoA) PDF for the given frameworks.

    Each requirement row lists the risks it addresses along with their
    residual level and treatment decision so auditors can trace control
    selection back to the risk that drove it.

    Returns a tuple (filename, content_bytes).
    """
    from weasyprint import HTML

    frameworks_data = build_soa_frameworks_data(frameworks)

    now = timezone.now()
    company = CompanySettings.get()
    html_string = render_to_string("reports/soa_pdf.html", {
        "frameworks_data": frameworks_data,
        "generated_at": now,
        "generated_by": user,
        "company": company,
    })

    pdf_bytes = HTML(string=html_string).write_pdf()

    date_str = now.strftime("%Y%m%d_%H%M%S")
    filename = f"SoA_{date_str}.pdf"

    return filename, pdf_bytes


def _build_scope_tree(scopes):
    """Build a tree structure from a flat list of scopes.

    Returns a list of dicts: {scope, depth, children} ordered depth-first.
    Each scope in the input is placed under its parent if the parent is also
    in the list; otherwise it becomes a root node.
    """
    scope_ids = {s.pk for s in scopes}
    scope_map = {s.pk: s for s in scopes}

    # Build children mapping
    children_map = {}  # parent_pk -> [child scopes]
    roots = []
    for s in scopes:
        parent_pk = s.parent_scope_id
        if parent_pk and parent_pk in scope_ids:
            children_map.setdefault(parent_pk, []).append(s)
        else:
            roots.append(s)

    # Sort roots and children by name
    roots.sort(key=lambda s: s.name)
    for k in children_map:
        children_map[k].sort(key=lambda s: s.name)

    # Flatten depth-first
    result = []

    def _walk(scope, depth):
        result.append({"scope": scope, "depth": depth})
        for child in children_map.get(scope.pk, []):
            _walk(child, depth + 1)

    for root in roots:
        _walk(root, 0)

    return result


def _compliance_color(status_key):
    """Return a CSS row color class for a compliance status."""
    if status_key in ("compliant", "strength"):
        return "row-compliant"
    if status_key in ("minor_non_conformity", "observation", "improvement_opportunity"):
        return "row-partial"
    if status_key == "major_non_conformity":
        return "row-non-compliant"
    if status_key == "not_applicable":
        return "row-na"
    return ""


def _finding_color(finding_type):
    """Return a CSS row color class for a finding type."""
    if finding_type == "strength":
        return "row-compliant"
    if finding_type in ("minor_nc", "observation", "improvement"):
        return "row-partial"
    if finding_type == "major_nc":
        return "row-non-compliant"
    return ""


def generate_audit_report_pdf(assessment, user):
    """Generate an Audit Report PDF for a completed compliance assessment.

    Returns a tuple (filename, content_bytes).
    """
    from weasyprint import HTML

    from compliance.constants import ComplianceStatus, FindingType

    # Gather results grouped by framework
    frameworks = list(assessment.frameworks.all())
    scopes = list(assessment.scopes.select_related("parent_scope"))
    scope_tree = _build_scope_tree(scopes)
    results = list(
        assessment.results
        .select_related("requirement__section", "requirement__framework", "assessed_by")
        .prefetch_related("attachments")
        .order_by("requirement__framework__name", "requirement__requirement_number")
    )

    # Build findings data
    findings = list(
        assessment.findings
        .prefetch_related("requirements")
        .select_related("assessor")
        .order_by("reference")
    )

    # Build requirement_id → finding mapping for per-framework finding cards
    req_id_to_findings = {}
    for f in findings:
        for req in f.requirements.all():
            req_id_to_findings.setdefault(req.pk, []).append(f)

    # Build findings recap (section 2.2)
    findings_recap = []
    for f in findings:
        reqs = list(f.requirements.all())
        req_display = ", ".join(
            r.requirement_number or r.reference for r in reqs
        )
        req_names = "; ".join(r.name for r in reqs)
        findings_recap.append({
            "reference": f.reference,
            "requirement_refs": req_display,
            "requirement_names": req_names,
            "type_display": f.get_finding_type_display(),
            "finding_type": f.finding_type,
            "color": _finding_color(f.finding_type),
        })

    # Build frameworks data
    frameworks_data = []
    for idx, fw in enumerate(frameworks):
        fw_results = [r for r in results if r.requirement.framework_id == fw.pk]

        # Compliance state rows (section X.1)
        compliance_rows = []
        for r in fw_results:
            compliance_rows.append({
                "number": r.requirement.requirement_number or r.requirement.reference,
                "name": r.requirement.name,
                "status": r.get_compliance_status_display(),
                "status_key": r.compliance_status,
                "color": _compliance_color(r.compliance_status),
            })
        compliance_rows.sort(key=lambda row: _natural_sort_key(row["number"]))

        # Findings detail cards for this framework (section X.2)
        fw_req_ids = {r.requirement.pk for r in fw_results}
        fw_findings = []
        seen_finding_ids = set()
        for f in findings:
            if f.pk in seen_finding_ids:
                continue
            f_req_ids = {r.pk for r in f.requirements.all()}
            if f_req_ids & fw_req_ids:
                seen_finding_ids.add(f.pk)
                reqs = [r for r in f.requirements.all() if r.pk in fw_req_ids]
                req_display = ", ".join(
                    r.requirement_number or r.reference for r in reqs
                )
                req_items = [{
                    "number": r.requirement_number or r.reference,
                    "name": r.name,
                    "description": _clean_html(r.description),
                } for r in reqs]
                fw_findings.append({
                    "reference": f.reference,
                    "type_display": f.get_finding_type_display(),
                    "finding_type": f.finding_type,
                    "color": _finding_color(f.finding_type),
                    "requirement_refs": req_display,
                    "requirement_items": req_items,
                    "description": _clean_html(f.description),
                    "recommendation": _clean_html(f.recommendation),
                    "evidence": _clean_html(f.evidence),
                })

        frameworks_data.append({
            "framework": fw,
            "section_num": idx + 3,
            "compliance_rows": compliance_rows,
            "findings": fw_findings,
        })

    # Summary statistics
    from collections import Counter
    finding_type_counts = Counter(f.finding_type for f in findings)
    summary = {
        "total": assessment.total_requirements,
        "overall_compliance": assessment.overall_compliance_level,
        "coverage_pct": assessment.coverage_pct,
        "total_findings": len(findings),
    }
    # Count distinct requirements impacted per finding type
    req_ids_by_type = {}
    for f in findings:
        ft = f.finding_type
        if ft not in req_ids_by_type:
            req_ids_by_type[ft] = set()
        for req in f.requirements.all():
            req_ids_by_type[ft].add(req.pk)

    total_impacted_reqs = len({
        req_pk for ids in req_ids_by_type.values() for req_pk in ids
    })

    _type_defs = [
        ("major_nc", FindingType.MAJOR_NON_CONFORMITY.label, "row-non-compliant"),
        ("minor_nc", FindingType.MINOR_NON_CONFORMITY.label, "row-partial"),
        ("observation", FindingType.OBSERVATION.label, "row-partial"),
        ("improvement", FindingType.IMPROVEMENT_OPPORTUNITY.label, "row-partial"),
        ("strength", FindingType.STRENGTH.label, "row-compliant"),
    ]

    findings_by_type = []
    for key, label, color in _type_defs:
        findings_by_type.append({
            "label": label,
            "finding_count": finding_type_counts.get(key, 0),
            "req_count": len(req_ids_by_type.get(key, set())),
            "color": color,
        })

    summary["total_impacted_reqs"] = total_impacted_reqs

    # Finding type definitions for section 1.4
    finding_qualifications = [
        {
            "type": FindingType.MAJOR_NON_CONFORMITY.label,
            "color": "row-non-compliant",
        },
        {
            "type": FindingType.MINOR_NON_CONFORMITY.label,
            "color": "row-partial",
        },
        {
            "type": FindingType.OBSERVATION.label,
            "color": "row-partial",
        },
        {
            "type": FindingType.IMPROVEMENT_OPPORTUNITY.label,
            "color": "row-partial",
        },
        {
            "type": FindingType.STRENGTH.label,
            "color": "row-compliant",
        },
    ]

    # Annexes: list of all analyzed documents (attachments)
    annexes_docs = []
    for r in results:
        for att in r.attachments.all():
            annexes_docs.append({
                "filename": att.original_filename,
                "requirement_ref": r.requirement.requirement_number or r.requirement.reference,
                "requirement_name": r.requirement.name,
                "uploaded_at": att.uploaded_at,
            })
    annexes_docs.sort(key=lambda d: (d["requirement_ref"], d["filename"]))

    now = timezone.now()
    company = CompanySettings.get()
    html_string = render_to_string("reports/audit_report_pdf.html", {
        "assessment": assessment,
        "company": company,
        "assessment_description": _clean_html(assessment.description),
        "assessment_limitations": _clean_html(assessment.limitations),
        "scope_tree": scope_tree,
        "frameworks": frameworks,
        "frameworks_data": frameworks_data,
        "findings_recap": findings_recap,
        "summary": summary,
        "findings_by_type": findings_by_type,
        "finding_qualifications": finding_qualifications,
        "annexes_docs": annexes_docs,
        "generated_at": now,
        "generated_by": user,
    })

    pdf_bytes = HTML(string=html_string).write_pdf()

    date_str = now.strftime("%Y%m%d_%H%M%S")
    filename = f"Audit_Report_{assessment.reference}_{date_str}.pdf"

    return filename, pdf_bytes


# ── Risk register XLSX ────────────────────────────────────────


RISK_REGISTER_COLUMNS = [
    "Reference",
    "Name",
    "Assessment",
    "Source",
    "Threats",
    "Vulnerabilities",
    "Essential assets",
    "Support assets",
    "Linked requirements",
    "Initial likelihood",
    "Initial impact",
    "Initial level",
    "Current likelihood",
    "Current impact",
    "Current level",
    "Residual likelihood",
    "Residual impact",
    "Residual level",
    "Treatment decision",
    "Treatment plans",
    "Owner",
    "Priority",
    "Status",
    "Review date",
    "Approved",
    "Created at",
]


def _risk_row(risk):
    """Return a list of cell values matching RISK_REGISTER_COLUMNS for one Risk."""
    threats = ", ".join(
        sorted({src.threat.name for src in risk.iso27005_sources.all() if src.threat_id})
    )
    vulnerabilities = ", ".join(
        sorted({src.vulnerability.name for src in risk.iso27005_sources.all() if src.vulnerability_id})
    )
    essential = ", ".join(sorted(a.name for a in risk.affected_essential_assets.all()))
    support = ", ".join(sorted(a.name for a in risk.affected_support_assets.all()))
    requirements = ", ".join(
        sorted(
            f"{r.requirement_number or r.reference}"
            for r in risk.linked_requirements.all()
        )
    )
    treatment_plans = ", ".join(
        sorted(f"{tp.reference} {tp.name}" for tp in risk.treatment_plans.all())
    )
    owner_name = ""
    if risk.risk_owner_id:
        owner = risk.risk_owner
        owner_name = (
            getattr(owner, "display_name", "")
            or getattr(owner, "email", "")
            or str(owner)
        )
    return [
        risk.reference or "",
        risk.name or "",
        str(risk.assessment) if risk.assessment_id else "",
        risk.get_risk_source_display(),
        threats,
        vulnerabilities,
        essential,
        support,
        requirements,
        risk.initial_likelihood,
        risk.initial_impact,
        risk.initial_risk_level,
        risk.current_likelihood,
        risk.current_impact,
        risk.current_risk_level,
        risk.residual_likelihood,
        risk.residual_impact,
        risk.residual_risk_level,
        risk.get_treatment_decision_display(),
        treatment_plans,
        owner_name,
        risk.get_priority_display(),
        risk.get_status_display(),
        risk.review_date.isoformat() if risk.review_date else "",
        "yes" if risk.is_approved else "no",
        risk.created_at.replace(tzinfo=None).isoformat(timespec="seconds")
        if risk.created_at
        else "",
    ]


def generate_risk_register_xlsx(risks_qs, user):
    """Generate an Excel workbook listing the risks in `risks_qs`.

    Returns a tuple (filename, content_bytes). The caller is responsible for
    filtering `risks_qs` by scope and any other criteria before calling; the
    lifecycle rule (only reportable elements appear in reports) is applied here.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    from core.workflow import reportable

    risks_qs = reportable(risks_qs)

    now = timezone.now()
    org_name = ""
    try:
        org_name = CompanySettings.get().name or ""
    except Exception:
        org_name = ""

    wb = Workbook()
    ws = wb.active
    ws.title = "Risk register"

    # Title row
    ws.append([f"Risk register - {org_name}".strip(" -")])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(RISK_REGISTER_COLUMNS))
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # Generation metadata
    generated_by = ""
    if user:
        generated_by = getattr(user, "display_name", "") or getattr(user, "email", "")
    ws.append([
        f"Generated at {now.strftime('%Y-%m-%d %H:%M')} - by {generated_by}".strip(" -"),
    ])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(RISK_REGISTER_COLUMNS))
    meta_cell = ws.cell(row=2, column=1)
    meta_cell.font = Font(italic=True, color="666666")
    meta_cell.alignment = Alignment(horizontal="center")

    ws.append([])  # spacer

    # Header row
    header_row_idx = ws.max_row + 1
    ws.append(RISK_REGISTER_COLUMNS)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx in range(1, len(RISK_REGISTER_COLUMNS) + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    risks_qs = (
        risks_qs.select_related("assessment", "risk_owner")
        .prefetch_related(
            "affected_essential_assets",
            "affected_support_assets",
            "linked_requirements",
            "iso27005_sources__threat",
            "iso27005_sources__vulnerability",
            "treatment_plans",
        )
        .order_by("reference")
    )
    for risk in risks_qs:
        ws.append(_risk_row(risk))

    # Auto-size columns based on header length (no per-cell width measurement)
    for col_idx, header in enumerate(RISK_REGISTER_COLUMNS, start=1):
        width = max(12, min(40, len(header) + 4))
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = ws.cell(row=header_row_idx + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    date_str = now.strftime("%Y%m%d_%H%M%S")
    filename = f"Risk_register_{date_str}.xlsx"
    return filename, buf.getvalue()
