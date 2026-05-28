"""Tests for the risk register XLSX export generator."""

import io
import zipfile

import pytest

from accounts.tests.factories import UserFactory
from compliance.tests.factories import FrameworkFactory, RequirementFactory
from context.tests.factories import ScopeFactory
from reports.generators import (
    RISK_REGISTER_COLUMNS,
    generate_risk_register_xlsx,
)
from risks.models import Risk
from risks.tests.factories import RiskAssessmentFactory, RiskFactory


pytestmark = pytest.mark.django_db


def _read_workbook(content):
    from openpyxl import load_workbook
    return load_workbook(io.BytesIO(content), read_only=False)


class TestRiskRegisterXlsx:
    def test_empty_queryset_produces_valid_xlsx(self):
        user = UserFactory(is_superuser=True)
        filename, content = generate_risk_register_xlsx(Risk.objects.none(), user)
        assert filename.endswith(".xlsx")
        # XLSX is a ZIP archive.
        zf = zipfile.ZipFile(io.BytesIO(content))
        assert "xl/workbook.xml" in zf.namelist()

    def test_header_row_matches_columns(self):
        user = UserFactory(is_superuser=True)
        _, content = generate_risk_register_xlsx(Risk.objects.none(), user)
        wb = _read_workbook(content)
        ws = wb.active
        # Title row 1, metadata row 2, spacer 3, header row 4.
        header_values = [c.value for c in ws[4]]
        assert header_values == RISK_REGISTER_COLUMNS

    def test_data_row_contains_risk_fields(self):
        user = UserFactory(is_superuser=True)
        assessment = RiskAssessmentFactory(name="Annual 2026")
        risk = RiskFactory(
            assessment=assessment,
            name="Data leak",
            current_likelihood=3,
            current_impact=4,
        )
        _, content = generate_risk_register_xlsx(
            Risk.objects.filter(pk=risk.pk), user,
        )
        wb = _read_workbook(content)
        ws = wb.active
        # Data starts at row 5
        row_values = [c.value for c in ws[5]]
        idx = {col: i for i, col in enumerate(RISK_REGISTER_COLUMNS)}
        assert row_values[idx["Reference"]] == risk.reference
        assert row_values[idx["Name"]] == "Data leak"
        assert row_values[idx["Current likelihood"]] == 3
        assert row_values[idx["Current impact"]] == 4

    def test_includes_linked_requirements(self):
        user = UserFactory(is_superuser=True)
        risk = RiskFactory()
        framework = FrameworkFactory()
        req1 = RequirementFactory(framework=framework, requirement_number="A.5.1")
        req2 = RequirementFactory(framework=framework, requirement_number="A.5.2")
        risk.linked_requirements.add(req1, req2)
        _, content = generate_risk_register_xlsx(
            Risk.objects.filter(pk=risk.pk), user,
        )
        wb = _read_workbook(content)
        ws = wb.active
        row_values = [c.value for c in ws[5]]
        idx = RISK_REGISTER_COLUMNS.index("Linked requirements")
        assert "A.5.1" in row_values[idx]
        assert "A.5.2" in row_values[idx]


class TestRiskRegisterExportView:
    def test_login_required(self):
        from django.test import Client
        from django.urls import reverse
        resp = Client().get(reverse("risks:risk-register-export-xlsx"))
        assert resp.status_code == 302

    def test_superuser_can_export(self):
        from django.test import Client
        from django.urls import reverse
        user = UserFactory(is_superuser=True, is_staff=True)
        client = Client()
        client.force_login(user)
        RiskFactory.create_batch(2)
        resp = client.get(reverse("risks:risk-register-export-xlsx"))
        assert resp.status_code == 200
        assert resp["Content-Type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert resp["Content-Disposition"].startswith("attachment;")
        # The export is persisted as a Report.
        from reports.constants import ReportType
        from reports.models import Report
        assert Report.objects.filter(report_type=ReportType.RISK_REGISTER).exists()

    def test_scope_filter_applied(self):
        from django.test import Client
        from django.urls import reverse
        scope_in = ScopeFactory()
        scope_out = ScopeFactory()

        # A risk inside the user's scope, one outside.
        assessment_in = RiskAssessmentFactory()
        assessment_in.scopes.add(scope_in)
        assessment_out = RiskAssessmentFactory()
        assessment_out.scopes.add(scope_out)
        risk_in = RiskFactory(assessment=assessment_in, name="Inside")
        risk_out = RiskFactory(assessment=assessment_out, name="Outside")

        from accounts.models import Group, Permission
        group = Group.objects.create(name="Test scope group")
        group.allowed_scopes.add(scope_in)
        for codename in ["risks.risk.read", "risks.export.read"]:
            perm, _ = Permission.objects.get_or_create(
                codename=codename,
                defaults={
                    "name": codename, "module": codename.split(".")[0],
                    "feature": codename.split(".")[1],
                    "action": codename.split(".")[2],
                    "is_system": True,
                },
            )
            group.permissions.add(perm)

        user = UserFactory(is_superuser=False, is_staff=False)
        group.users.add(user)

        client = Client()
        client.force_login(user)
        resp = client.get(reverse("risks:risk-register-export-xlsx"))
        assert resp.status_code == 200
        wb = _read_workbook(resp.content)
        ws = wb.active
        all_text = "\n".join(
            " ".join(str(c.value) for c in row if c.value)
            for row in ws.iter_rows(min_row=5)
        )
        assert "Inside" in all_text
        assert "Outside" not in all_text

    def test_query_filters_applied(self):
        from django.test import Client
        from django.urls import reverse
        user = UserFactory(is_superuser=True, is_staff=True)
        client = Client()
        client.force_login(user)
        RiskFactory(name="StatusA", status="analyzed")
        RiskFactory(name="StatusB", status="closed")
        resp = client.get(
            reverse("risks:risk-register-export-xlsx") + "?status=closed",
        )
        assert resp.status_code == 200
        wb = _read_workbook(resp.content)
        ws = wb.active
        all_text = "\n".join(
            " ".join(str(c.value) for c in row if c.value)
            for row in ws.iter_rows(min_row=5)
        )
        assert "StatusB" in all_text
        assert "StatusA" not in all_text
